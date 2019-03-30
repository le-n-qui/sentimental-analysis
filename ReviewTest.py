#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 15 09:32:37 2018

@author: Kriti
"""

"""Importing packages"""
import nltk
import pickle
import openpyxl
from openpyxl import load_workbook
import os
import math
import xlsxwriter

#File to store predictions
results = xlsxwriter.Workbook('results.xlsx')
results_sheet = results.add_worksheet()

#Loading test set
test = 'rmp_test.xlsx'
wb = load_workbook(test)
ws = wb.active

# restore model from reviews.nb
with open('reviews.nb', 'rb') as f:
	model = pickle.load(f)

#total number of cases 
total = model['pos_count'] + model['neg_count']

#P(pos)
p_pos = model['pos_count']/ total

#P(neg)
p_neg = model['neg_count']/ total

pos_n = model['pos_fd'].N()
neg_n = model['neg_fd'].N()

fdist = model['pos_fd'] + model['neg_fd']
vocab_size = fdist.B()

for row in ws.iter_rows(min_col=1, max_col= 10, min_row = 2, max_row=14445):
	for cell in row:
		if(cell.column == 'C'):
			professor = cell.value
		
		if(cell.column == 'E'):
			review = ws.cell(row = cell.row, column = 5)
			file = open("Reviews/" + professor, "w")
			if(review.value != None):
				file.write(review.value)

"""Actual Predictions Implementing Model"""				
test_corpus = nltk.corpus.PlaintextCorpusReader("Reviews", '.*') 			
files = test_corpus.fileids()
row_count = 0

for f in files:	
	log_cpos = 0
	log_cneg = 0
	wordlist = test_corpus.words(f)
	
	for word in wordlist:
		posp= ((model['pos_fd'][word] + 1)/(pos_n + vocab_size))	
		log_pos = math.log(posp) 
		log_cpos = log_cpos + log_pos 
		negp= ((model['neg_fd'][word] + 1)/(neg_n + vocab_size))    
		log_neg= math.log(negp)
		log_cneg = log_cneg + log_neg
	p_num = (log_cpos + math.log(p_pos))  
	hp_num = (log_cneg + math.log(p_neg))  
	p_denom = p_num + hp_num


	pos_in_msg = p_num / p_denom 
	neg_in_msg = hp_num / p_denom 

	col = 2
	results_sheet.write(row_count, 1, f)
	if(pos_in_msg < neg_in_msg): 
		results_sheet.write(row_count, col, "Positive")
	else:
		results_sheet.write(row_count, col, "Negative")		
	row_count = row_count + 1
results.close()

	


			


