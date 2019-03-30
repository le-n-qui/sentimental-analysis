#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 15 09:32:37 2018

@author: Kriti
"""

"""Importing packages"""
import nltk
import pickle
import openpyxl
from openpyxl import load_workbook
import os

"""Opening training set"""
training = 'rmp_dump.xlsx'
pos_set  = open('pos_train.txt', 'w')
neg_set  = open('neg_train.txt', 'w')

wb = load_workbook(training)

ws = wb.active

pos_count = 0 #number of positive reviews
neg_count = 0 #number of negative reviews

"""Iterating through rows in training set""""
for row in ws.iter_rows(min_row=1, max_col= 10, max_row=59293):
	for cell in row:
		#print(row)
		if(cell.value == 'pos'): #If review is annotated as positive
			pos = ws.cell(row = cell.row, column = 5) 
			pos_count = pos_count + 1			
			pos_set.write(pos.value) #Save positive review in positive corpus

		if(cell.value == 'neg'): #If review is annotated as negative
			neg = ws.cell(row = cell.row, column = 5)
			neg_count = neg_count + 1			
			if(neg.value != None):
				neg_set.write(neg.value) #Save negative review in negative corpus

#Closing files
pos_set.close()
neg_set.close()


"""Opening positive reviews and negative reviews into copora"""
pos_corpus = nltk.corpus.PlaintextCorpusReader(".", "pos_train.txt") 
neg_corpus = nltk.corpus.PlaintextCorpusReader(".", "neg_train.txt")

"""Applying FreqDist() to words in positive corpus and negative corpus"""
pos_words = pos_corpus.words()
pos_fd = nltk.FreqDist(pos_words)

neg_words = neg_corpus.words()
neg_fd = nltk.FreqDist(neg_words)

model = {
	'pos_count' : pos_count,
	'neg_count' : neg_count,
	'pos_fd' : pos_fd,
	'neg_fd' : neg_fd
}

"""Using pickle to save file"""
pickle.dump(model, open('reviews.nb', 'wb'))

        